from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
import time
import pyperclip
import pandas as pd
import subprocess
from datetime import datetime
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import glob
import os
import re

# 사용자가 환경에 따라 변경해야 할 값
user = 'movem'
ver = str("2023-07-04")
auth_dic = {'id':'pw'}
chrome_ver = 114
filter_list = ['사이트명', '사이트주소', '사용아이디', '업로드여부', '파일명']

# 크롬드라이버로

subprocess.Popen(
    r'C:\Program Files\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\\Users\\' + user + r'\\AppData\\Local\\Google\\Chrome\\User Data"')
option = Options()
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = webdriver.Chrome(options=option)
driver.execute_script('window.open("about:blank", "_blank");')
tabs = driver.window_handles

# 안내
print("\n")
print("씽굿 프로그램입니다.")
print("https://github.com/movemin03/CafePosting_Auto")
print("ver:" + ver)

# 정보 입력
print('필요한 정보를 기입해야합니다\n')
blank_auth_dic = {}
slash_auth_dic = {}
# 슬레쉬 제거
for key, value in auth_dic.items():
    left_key = key.split('/')[0]
    slash_auth_dic[left_key] = value
auth_dic.update(slash_auth_dic)
# 공백 제거
for key, value in auth_dic.items():
    new_key = key.strip()
    blank_auth_dic[new_key] = value
auth_dic.update(blank_auth_dic)
print("게시물의 제목을 적어주십시오:")
title = input()
print("복사할 html 코드 위치를 입력해주세요:")
content_path = input().replace('"', '')

# 데이터 전처리0
print('참고할 엑셀 파일 위치를 알려주세요:')
upload_path = input().replace('"', '')
print('\n 입력하신 엑셀파일을 읽어오고 있습니다')
excel = pd.read_excel(upload_path, names=['사이트명', '사이트주소', '사용아이디', '업로드여부', '파일명'])
input_id_list = list(excel['사용아이디'].drop_duplicates())
try:
    input_id_list = [item.strip() for item in input_id_list]
    input_id_list = [string.split("/")[0].strip() for string in input_id_list]
except:
    print("100 퍼센트의 확률로 제공하신 엑셀파일의 아이디 칸이 비어 있습니다. 프로그램을 재실행하시길 권장합니다")

n_error_list, d_error_list = [], []
wait = WebDriverWait(driver, 5)
error_posting_url = 0


def content_html():
    driver.switch_to.window(tabs[0])
    driver.get(content_path)
    action = ActionChains(driver)
    action.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
    action.key_down(Keys.CONTROL).send_keys('c').key_up(Keys.CONTROL).perform()
    driver.switch_to.window(tabs[1])

def login():
    global tabs
    tabs = driver.window_handles
    try:
        if len(tabs) >= 2:
            driver.switch_to.window(tabs[1])
        else:
            print("탭이 부족해서 새 탭을 엽니다")
            driver.switch_to.window(tabs[0])
            driver.execute_script('window.open("about:blank", "_blank");')
            tabs = driver.window_handles
            driver.switch_to.window(tabs[1])
            time.sleep(2)
    except:
        print("탭 갯수: " + str(len(tabs)))
        print("탭 관리에 문제가 있습니다")
        print("탭을 수동으로 열어주십시오")
        a = input()

    login_url = "https://nid.naver.com/nidlogin.login"
    driver.get(login_url)
    print("로그인 링크 접속 완료: " + login_url)

    pyperclip.copy(auth)
    driver.find_element(By.ID, 'id').click()
    action = ActionChains(driver)
    action.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
    action.send_keys(Keys.BACKSPACE).perform()
    driver.find_element(By.ID, 'id').send_keys(Keys.CONTROL + 'v')

    try:
        password = auth_dic[auth]
    except:
        print("auth_dic을 가져오는 것에 오류가 있어서 패스워드를 수동입력해야 합니다:")
        password = input()
    pyperclip.copy(password)
    driver.find_element(By.ID, 'pw').click()
    action.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
    action.send_keys(Keys.BACKSPACE).perform()
    driver.find_element(By.ID, 'pw').send_keys(Keys.CONTROL + 'v')

    time.sleep(1)
    login_btn = driver.find_element(By.ID, 'log.login')
    login_btn.click()
    print("로그인: 로그인 작업 진행 완료\n")
    a = input("2차 인증 여부 및 아이디가 " + auth + "가 맞는지 확인해주시고 아무거나 입력 후 엔터")


def login_daum():
    global tabs
    tabs = driver.window_handles
    try:
        if len(tabs) >= 2:
            driver.switch_to.window(tabs[1])
        else:
            print("탭이 부족해서 새 탭을 엽니다")
            driver.switch_to.window(tabs[0])
            driver.execute_script('window.open("about:blank", "_blank");')
            tabs = driver.window_handles
            driver.switch_to.window(tabs[1])
            time.sleep(2)
    except:
        print("탭 갯수: " + str(len(tabs)))
        print("탭 관리에 문제가 있습니다")
        print("탭을 수동으로 열어주십시오")
        a = input()

    login_url = "https://logins.daum.net/accounts/ksso.do?url=https%3A%2F%2Fwww.daum.net"
    driver.get(login_url)
    print("로그인 링크 접속 완료: " + login_url)

    time.sleep(1)
    pyperclip.copy(auth)
    action = ActionChains(driver)
    driver.find_element(By.ID, 'password--1').click()
    action.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
    action.send_keys(Keys.BACKSPACE).perform()
    driver.find_element(By.ID, 'loginKey--1').send_keys(Keys.CONTROL + 'v')

    try:
        password = auth_dic[auth]
    except:
        print("auth_dic을 가져오는 것에 오류가 있어서 패스워드를 수동입력해야 합니다:")
        password = input()
    pyperclip.copy(password)
    driver.find_element(By.ID, 'password--2').click()
    action.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
    action.send_keys(Keys.BACKSPACE).perform()
    driver.find_element(By.ID, 'password--2').send_keys(Keys.CONTROL + 'v')
    time.sleep(1)
    login_btn = driver.find_element(By.CLASS_NAME, 'confirm_btn').find_element(By.CLASS_NAME, 'submit')
    login_btn.click()
    a = input("2차 인증 여부 및 아이디가 " + auth + "가 맞는지 확인해주시고 아무거나 입력 후 엔터")


def posting():
    global tabs
    tabs = driver.window_handles
    try:
        if len(tabs) >= 2:
            driver.switch_to.window(tabs[1])
        else:
            print("탭이 부족해서 새 탭을 엽니다")
            driver.switch_to.window(tabs[0])
            driver.execute_script('window.open("about:blank", "_blank");')
            tabs = driver.window_handles
            driver.switch_to.window(tabs[1])
            time.sleep(2)
    except:
        print("탭 갯수: " + str(len(tabs)))
        print("탭 관리에 문제가 있습니다")
        print("탭을 수동으로 열어주십시오")
        a = input()

    driver.get(naver_url)
    print("링크 접속 완료: " + naver_url)

    error_myactivity = 0
    # 포스팅
    try:
        driver.find_element(By.CLASS_NAME, 'cafe-info-action').find_element(By.XPATH,
                                                                            '//a[contains(text(),"나의활동")]').click()
        wait.until(EC.presence_of_element_located((By.XPATH, '//a[contains(text(),"내가 쓴 글 보기")]')))
        driver.find_element(By.XPATH, '//a[contains(text(),"내가 쓴 글 보기")]').click()
    except:
        try:
            driver.find_element(By.CLASS_NAME, 'cafe-info-action').find_element(By.XPATH,
                                                                                '//a[contains(text(),"나의활동")]').click()
            wait.until(EC.presence_of_element_located((By.XPATH, '//a[contains(text(),"내가 쓴 글 보기")]')))
            driver.find_element(By.XPATH, '//a[contains(text(),"내가 쓴 글 보기")]').click()
        except:
            print(
                '오류: 가입되지 않은 카페 or 강퇴 or (낮은 확률로 활동정지)에 의한 오류입니다. 아이디가 ' + auth + '가 맞는지 확인하고 아니라면 재로그인해주세요. 이후 아무키나 눌러주십시오')
            a = input()
            try:
                driver.find_element(By.CLASS_NAME, 'cafe-info-action').find_element(By.XPATH,
                                                                                    '//a[contains(text(),"나의활동")]').click()
                wait.until(EC.presence_of_element_located((By.XPATH, '//a[contains(text(),"내가 쓴 글 보기")]')))
                driver.find_element(By.XPATH, '//a[contains(text(),"내가 쓴 글 보기")]').click()
            except:
                print('오류: 가입되지 않은 카페 or 강퇴 or (낮은 확률로 활동정지)에 의한 오류')
                error_myactivity = 1

    # frame으로 변환해야 게시글 확인이 가능하다#

    if error_myactivity == 0:
        time.sleep(1)
        former_post = 1
        try:
            driver.switch_to.frame("cafe_main")
            mcn = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/table/tbody/tr[1]/td[1]/div[3]/div/a')
            mcn_lnk = mcn.get_attribute('href')
            driver.get(mcn_lnk)

        except:
            if error_myactivity == 0:
                print('이전 게시물이 없어서 프로그램이 참조할 것이 없습니다.')
                former_post = 0
            else:
                pass

        driver.switch_to.default_content()
        wait.until(EC.presence_of_element_located((By.NAME, "cafe_main")))
        driver.switch_to.frame("cafe_main")
        # 이전 포스트가 있는 경우 이전 포스트 내 글쓰기 누르고, 없는 경우정지상태 새 글 작성
        global status_stop
        status_stop = 0
        if former_post == 1:
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div/div[3]/div[1]/a[1]')))
                mcn = driver.find_element(By.XPATH, '//*[@id="app"]/div/div/div[3]/div[1]/a[1]')
                mcn_lnk = mcn.get_attribute('href')
                driver.get(mcn_lnk)
            except:
                print("높은 확률로 정지 상태입니다")
                status_stop = 1
        else:
            wait.until(EC.presence_of_element_located((By.XPATH, '//span[contains(@class,"BaseButton__txt")]')))
            mcn = driver.find_element(By.XPATH, '//span[contains(@class,"BaseButton__txt")]').find_element(By.XPATH,
                                                                                                           './..')
            mcn_lnk = mcn.get_attribute('href')
            driver.get(mcn_lnk)

        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, '//*[@id="app"]/div/div/section/div/div[2]/div[1]/div[1]/div[2]/div/textarea')))
            driver.find_element(By.XPATH,
                                '//*[@id="app"]/div/div/section/div/div[2]/div[1]/div[1]/div[2]/div/textarea').click()

            action = ActionChains(driver)
            action.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            action.send_keys(title).perform()
            print("글쓰기 1/3: 제목 입력 완료")

            content_html()
            wait.until(EC.presence_of_element_located(
                (By.XPATH, '//p[contains(@class,"se-text-paragraph se-text-paragraph-align-left")]')))
            driver.find_elements(By.XPATH, '//p[contains(@class,"se-text-paragraph se-text-paragraph-align-left")]')[
                0].click()
            action.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            action.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            print("글쓰기 2/3: html 코드 입력 완료")

            if former_post == 0:
                print('카테고리를 수동 변경해야 합니다. 카테고리를 설정해주세요. 수정 후 아무키나 눌러주세요')
                a = input()
            else:
                time.sleep(3)

            driver.find_element(By.XPATH, '//span[contains(@class,"BaseButton__txt")]').click()
            print("글쓰기 3/3: 업로드 완료")

            try:
                alert = driver.switch_to.alert  # alert 창에 대한 참조를 가져옵니다.
                alert_text = alert.text  # alert 창에 표시된 텍스트를 가져옵니다.
                print(str(alert_text))
                alert.accept()
                time.sleep(4)
                driver.find_element(By.XPATH, '//button[contains(text(),"등록")]').click()
                print("글쓰기 3/3: 업로드 완료: 연속된 글쓰기로 감지되었음")
            except:
                pass

            time.sleep(1)
            try:
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'cafe_default')))
                global posting_url_n
                posting_url_n = "NaN"
                posting_url_n = str(driver.current_url)
                if "articles/write" in posting_url_n:
                    posting_url_n = f"잘못된 링크가 들어갔으므로 수동 작업 필요:write: {posting_url_n}"
                    error_posting_url = 1
                elif posting_url_n == "NaN":
                    posting_url_n = f"잘못된 링크가 들어갔으므로 수동 작업 필요:NaN: + {naver_url}"
                    error_posting_url = 1
                else:
                    pass
            except:
                posting_url_n = "(금칙어 처리) 혹은 연속된 게시물 또는 창이 꺼짐. 본래 url: " + naver_url
                error_posting_url = 1
            time.sleep(2)
        except:
            print("높은 확률로 정지 상태입니다")
            status_stop = 1
    else:
        pass

def posting_daum():
    global tabs
    tabs = driver.window_handles
    try:
        a = input("탭전환합니다")
        if len(tabs) >= 2:
            driver.switch_to.window(tabs[1])
        else:
            print("탭이 부족해서 새 탭을 엽니다")
            driver.switch_to.window(tabs[0])
            driver.execute_script('window.open("about:blank", "_blank");')
            tabs = driver.window_handles
            driver.switch_to.window(tabs[1])
            time.sleep(2)
    except:
        print("탭 갯수: " + str(len(tabs)))
        print("탭 관리에 문제가 있습니다")
        print("탭을 수동으로 열어주십시오")
        a = input()
    driver.get(daum_url)
    print("링크 접속 완료: " + daum_url)

    # 포스팅
    error_myactivity = 0
    time.sleep(1)
    try:
        driver.switch_to.default_content()
        wait.until(EC.presence_of_element_located((By.ID, "down")))
        driver.switch_to.frame("down")
        driver.find_element(By.XPATH, '//span[contains(text(),"내 정보")]').click()
        time.sleep(2)
        img_elem = driver.find_element(By.XPATH, '//*[@id="cafe_write_article_btn"]/img')
        parent_elem = img_elem.find_element(By.XPATH, './..')
        mcn = parent_elem.get_attribute('href')
        result = re.search(r'grpid=(.*?)&fldid', mcn)
        user_value = result.group(1)
        mcn_link = "https://cafe.daum.net/_c21_/member_article_cafesearch?item=userid&grpid=" + str(user_value)
        driver.get(mcn_link)
    except:
        print('오류: 가입되지 않은 카페 or 강퇴 or 활동정지에 의한 오류입니다. 아이디가 ' + auth + '가 맞는지 확인하고 아니라면 재로그인해주세요. 이후 아무키나 눌러주십시오')
        a = input()
        try:
            try:
                wait.until(EC.element_to_be_clickable((By.XPATH, '//a[contains(text(),"내가 쓴 글")]')))
                driver.find_element(By.XPATH, '//a[contains(text(),"내가 쓴 글")]').click()
            except:
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="myinfo_list"]/ul/li[3]/span[1]/a')))
                driver.find_element(By.XPATH, '//*[@id="myinfo_list"]/ul/li[3]/span[1]/a').click()
        except:
            print("내 정보 - 내가 쓴 글을 수동으로 눌러주십시오. 이후 이행했다면 0 이행하지 못한다면 1을 눌러주십시오:")
            while True:
                error_myactivity = int(input("0 또는 1을 입력해주세요: "))
                if error_myactivity == 0 or error_myactivity == 1:
                    break
                else:
                    print("잘못된 입력입니다. 다시 입력해주세요.")
    former_post = 1
    # 이전 게시글 입력창 접근
    if error_myactivity == 0:
        try:
            driver.switch_to.default_content()
            wait.until(EC.presence_of_element_located((By.ID, "down")))
            driver.switch_to.frame("down")

            time.sleep(2)
            mcn = driver.find_elements(By.CLASS_NAME, 'subject')[1].find_element(By.XPATH, './*[2]')
            mcn_lnk = mcn.get_attribute('href')
            driver.get(mcn_lnk)

            # 게시글 입력창 접근
            wait.until(EC.presence_of_element_located((By.ID, 'down')))
            driver.switch_to.frame("down")
            driver.find_element(By.ID, 'article-write-btn-bottom').click()
        except:
            print('이전 게시물이 없어서 프로그램이 참조할 것이 없습니다.')
            former_post = 0
            try:
                driver.find_element(By.XPATH, '//*[@id="cafe_write_article_btn"]/img').click()
            except:
                print("카페 글쓰기 버튼 클릭이 실패하여 수동으로 눌러주세요")
                a = input()
            # 게시글 입력창
        time.sleep(1)
        try:
            driver.find_element(By.CLASS_NAME, 'title__input').click()
        except:
            try:
                driver.find_element(By.CLASS_NAME, 'title__input').click()
            except:
                print('임시저장 글 원인으로 오류가 났습니다. 해결 후 엔터')
                a = input()

        driver.find_element(By.CLASS_NAME, 'title__input').click()
        action = ActionChains(driver)
        action.send_keys(title).perform()
        print("글쓰기 1/3: 제목 입력 완료")

        content_html()
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame("down")
            driver.switch_to.frame(driver.find_element(By.ID, 'keditorContainer_ifr'))
        except:
            print("iframe 바꾸기 실패")

        driver.find_element(By.XPATH, '//*[@id="tinymce"]/p').click()
        action.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
        print("글쓰기 2/3: html 코드 입력 완료")

        if former_post == 0:
            print('이전 게시글이 없어서 카테고리를 수동 조절해야 합니다. 수동 조절 후 아무거나 입력')
            a = input()
        else:
            time.sleep(3)
        driver.switch_to.default_content()
        driver.switch_to.frame("down")
        driver.find_element(By.XPATH, '//button[contains(text(),"등록")]').click()
        print("글쓰기 3/3: 업로드 완료")

        try:
            driver.switch_to.default_content()
            wait.until(EC.presence_of_element_located((By.ID, 'down')))
            driver.switch_to.frame("down")
            print("프레임 변환 완료")
            time.sleep(1)
            mcn = driver.find_element(By.CLASS_NAME, 'btn_center05').find_element(By.XPATH, "./*").get_attribute('href')
            driver.get(mcn)
            time.sleep(1)

            global posting_url_d
            posting_url_d = str(driver.current_url)
        except Exception as e:
            print("오류 발생:", e)
            posting_url_d = "업로드 링크를 가져오기 실패: 창 꺼짐, 포스팅 금지 등의 사유: " + daum_url
        time.sleep(1)
    else:
        pass


def combined():
    print('파일 병합을 하시겠습니까? y/n')
    a = input()
    if a == "y":
        combine_path = "C:\\Users\\" + user + "\\Desktop\\결과"
        print("병합된 파일은 여기에 저장됩니다: " + combine_path)
        print('위치 변경을 원하시나요? 변경 원한다면 y 입력')
        a = input()
        if a == "y":
            combine_path = input().replace('"', '')
            combine_path = combine_path + "\\"
        else:
            combine_path = combine_path + "\\"

        target = (str(combine_path) + '*.xlsx')
        print(target)
        xlsx_list = glob.glob(target)

        print("현재 있는 파일 :" + str(xlsx_list))

        dfs = []
        try:
            # 각 xlsx 파일에 대해 반복합니다.
            for xlsx_file in xlsx_list:
                # 엑셀 파일 읽기
                wb = load_workbook(xlsx_file)
                ws = wb.active
                row_heights = {}

                for row in range(1, ws.max_row + 1):
                    if row in ws.row_dimensions:
                        row_heights[row] = ws.row_dimensions[row].height
                    else:
                        row_heights[row] = "h_no"
                none_value_keys = [key for key, value in row_heights.items() if value is None]
                print("숨김 행:" + str(none_value_keys))

                excel = pd.read_excel(xlsx_file, header=0, skiprows=lambda x: x + 1 in none_value_keys)
                excel_re = excel[filter_list]
                print(excel_re)
                # 데이터프레임 리스트에 추가합니다.
                dfs.append(excel_re)

                # 데이터프레임을 모두 결합합니다.
                combined_df = pd.concat(dfs, ignore_index=True)

                # 새 엑셀 파일 저장하기
                output_file = os.path.join(combine_path, "combined.xlsx")
                combined_df.to_excel(output_file, index=False, header=True)
                print(f"합쳐진 파일이 {output_file}에 저장되었습니다.")
                error_code = 0
        except:
            print("형식이 다른 파일이 끼어있습니다. 찾아서 삭제해주세요. 다시 시도합니다: ")
            error_code = 1
    else:
        pass


# 실행되는 라인
print('본 프로그램에 등록되어 있는 id 는 다음과 같습니다:' + str(list(auth_dic.keys())))
if set(input_id_list) - set(list(auth_dic.keys())) == set():
    print('모두 자동 업로드 가능합니다')
    pre_list = str(input_id_list)
else:
    print('요청하신 id는 다음과 같습니다: ' + str(input_id_list))
    same = set(list(auth_dic.keys())) & set(input_id_list)
    if same == set():
        print('현재 자동 업로드를 진행할 수 있는 아이디가 없습니다')
        exit()
    else:
        print('요청하신 내용의 일부인 ' + str(same) + '를 자동 업로드할 수 있습니다')
        pre_list = str(same)
    print('auth_dic 에 다음 값을 추가해야 합니다 : ' + str(set(input_id_list) - set(list(auth_dic.keys()))))
    print('추가 없이 진행 시, 추가하지 않은 id 는 업로드하지 않습니다\n')
    a = input
# 불필요한 기호 생략
pre_list = pre_list.translate(str.maketrans("", "", "{}',"))

List = [x for x in pre_list.split()]
for x in List:
    auth = x
    auth = auth.replace("[", "")
    auth = auth.replace("]", "")
    auth = auth.replace("[", "")
    print("사용할 아이디: " + auth)
    execute_time = datetime.today().strftime("%Y%m%d_%H%M")

    # 데이터 전처리
    excel_1 = excel[excel['사용아이디'] == auth]
    print(excel_1)

    url_list = list(excel_1['사이트주소'])

    naver_list = [x for x in url_list if "cafe.naver.com" in x]
    len_naver = len(naver_list)
    daum_list = [y for y in url_list if "cafe.daum.net" in y]
    len_daum = len(daum_list)

    if len_naver > 0:
        try:
            login()
        except:
            print('오류: 이미 로그인이 진행 or 자동로그인 기능 겹침 or id/pw 오류. 수동 진행 후 아무 키 입력. 수동 진행 후 아무 키 입력')
            a = input()
    else:
        print('업로드할 naver 링크가 없습니다')
    print("\n" + auth + " 아이디로 네이버 카페 " + str(len_naver) + "개, 다음 카페 " + str(len_daum) + "개를 진행하겠습니다")
    i = 0
    while i < len_naver:
        try:
            naver_url = naver_list[i]
            posting()
            if error_posting_url == 1:
                excel_1.iloc[i, 3] = "확인필요"
                excel_1.iloc[i, 1] = posting_url_n
            else:
                excel_1.iloc[i, 3] = "O"
                excel_1.iloc[i, 1] = posting_url_n
            posting_url_n = "NaN"
        except:
            if i >= len_naver:
                pass
            else:
                n_error_list.append(naver_list[i])
                if status_stop == 1:
                    excel_1.iloc[i, 3] = "X-활동정지"
                else:
                    excel_1.iloc[i, 3] = "X"
        i = i + 1
    naver_list = []
    len_naver = 0

    if len_daum > 0:
        try:
            if "@" not in auth:
                auth += "@daum.net"
            else:
                pass
            login_daum()
        except:
            print('오류: 이미 로그인이 진행 or 자동로그인 기능 겹침 or id/pw 오류. 수동 진행 후 아무 키 입력')
            a = input()
    else:
        print('업로드할 daum 링크가 없습니다')
    ii = 0
    while ii < len_daum:
        try:
            daum_url = daum_list[ii]
            driver = webdriver.Chrome(options=option)
            posting_daum()
            excel_1.iloc[ii + i, 3] = "O"
            excel_1.iloc[ii + i, 1] = posting_url_d
        except:
            if ii >= len_daum:
                pass
            else:
                d_error_list.append(daum_list[ii])
                excel_1.iloc[ii + i, 3] = "X"
        ii = ii + 1

    daum_list = []
    len_daum = 0

    # 결과 폴더 경로 지정
    result_dir = 'C:\\Users\\' + user + '\\Desktop\\결과'

    # 결과 폴더가 없으면 생성
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
    excel_1.to_excel(result_dir + "\\" + auth + execute_time + '_작업완료.xlsx')

if len(n_error_list) > 0:
    print("다음은 권한이 없거나 오류가 있어서 업로드 하지 못한 링크들 입니다. 네이버 카페:" + str(n_error_list))
else:
    pass
if len(d_error_list) > 0:
    print("다음은 권한이 없거나 오류가 있어서 업로드 하지 못한 링크들 입니다. 다음 카페:" + str(d_error_list))
else:
    pass
print('작업완료된 내역을 엑셀 파일로 저장하였습니다.')

error_code = 0
while error_code == 0:
    combined()
