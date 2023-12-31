import glob
import os
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime

user = os.path.expanduser('~')
filter_list = ['사이트명', '사이트주소', '사용아이디', '업로드여부', '파일명']

def modify_category(category):
    indices = []
    num_list = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
    for i, char in enumerate(category):
        if char in num_list:
            indices.append((char, i))

    indices.sort(key=lambda x: x[1])  # 인덱스 위치를 기준으로 정렬

    number = ''.join([char for char, _ in indices])  # 정렬된 숫자 값을 합쳐서 number 변수에 할당
    first_num_index = indices[0][1]  # 숫자 시작 index
    korean = category[:first_num_index].replace(" ", "")  # 한글 부분
    category_modified = korean + " " + category[first_num_index:].replace(" ", "")

    return category_modified

def combined():
    print("결과 파일 병합 프로그램입니다")
    a = "y"
    if a == "y":
        combine_path = user + "\\Desktop\\결과"
        print("병합된 파일은 여기에 저장됩니다: " + combine_path)
        print('위치 변경을 원하시나요? 변경 원한다면 y 입력, 아니라면 아무거나 입력')
        a = input()
        if a == "y":
            combine_path = input().replace('"', '')
            combine_path = combine_path + "\\"
        else:
            combine_path = combine_path + "\\"

        target = (str(combine_path) + '*.xlsx')
        print(target)
        pre_xlsx_list = glob.glob(target)
        combined_file_exists = False
        xlsx_list = []
        for file in pre_xlsx_list:
            if "combined" in file.lower():
                combined_file_exists = True
            else:
                xlsx_list.append(file)
        print("현재 있는 파일 :" + str(xlsx_list))

        dfs = []
        try:
        # 각 xlsx 파일에 대해 반복합니다.
            for xlsx_file in xlsx_list:
            # 엑셀 파일 읽기
                wb = load_workbook(xlsx_file)
                ws = wb.active
                row_heights = {}

                for row in range(1, ws.max_row + 1):
                    if row in ws.row_dimensions:
                        row_heights[row] = ws.row_dimensions[row].height
                    else:
                        row_heights[row] = "h_no"
                none_value_keys = [key for key, value in row_heights.items() if value is None]
                if none_value_keys == []:
                    pass
                else:
                    print("숨김 행:" + str(none_value_keys))

                excel = pd.read_excel(xlsx_file, header=0, skiprows=lambda x: x+1 in none_value_keys)
                excel_re = excel[filter_list]
            # 데이터프레임 리스트에 추가합니다.
                dfs.append(excel_re)

            # 데이터프레임을 모두 결합합니다.
                combined_df = pd.concat(dfs, ignore_index=True)
            # 결합된 데이터프레임을 수정합니다
            # 결합된 데이터의 파일명에 공백이 제대로 적용되어 있는지 확인하고 변경합니다
                combined_df['파일명'] = combined_df['파일명'].apply(lambda x: modify_category(x) if any(char.isdigit() for char in x) else x)

            # 사이트명', '사이트주소', '사용아이디', '업로드여부', '파일명'
            # 1. 열 변경
                combined_df = combined_df.reindex(columns=['사이트명','사이트주소', '업로드여부','사용아이디','파일명'])
            # 2. 텍스트 나누고 카테고리명과 순번 오름차순 정렬
                try:
                    combined_df[['카테고리명', '순번']] = combined_df['파일명'].str.split(' ', expand=True)
                    del combined_df['파일명']
                    combined_df['순번'] = combined_df['순번'].astype(int)
                    combined_df = combined_df.sort_values(by=['카테고리명', '순번'], ascending=[True, True])
                    category_error = False
                except Exception as e:
                    if str(e) == "int() argument must be a string, a bytes-like object or a real number, not 'NoneType'":
                        category_error = True
                    else:
                        print("오류 발생:", str(e))
                        category_error = True

            # 새 엑셀 파일 저장하기
                now = datetime.now()
                formatted_datetime = now.strftime("_%y%m%d_%H%M")
                output_file = os.path.join(combine_path, "combined" + formatted_datetime + ".xlsx")

                if category_error == False:
                    categories = combined_df['카테고리명'].unique()
                    with pd.ExcelWriter(output_file) as writer:
                        for category in categories:
                            filtered_df = combined_df[combined_df['카테고리명'] == category]
                            sheet_name = category if category else "Uncategorized"
                            filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    pass
            if category_error == True:
                print("파일명(카테고리명)항목에 띄어쓰기가 누락되어 있습니다. 카테고리명과 숫자 사이 띄어쓰기를 추가해주십시오 ex)기본커뮤니티 20")
                print("일부 파일명 항목에 띄어쓰기가 제대로 되지 않아서, 카테고리별 시트 분리 기능과 정렬 기능을 적용하지 않습니다")
                combined_df.to_excel(output_file, index=False, header=True)

            print(f"합쳐진 파일이 {output_file}에 저장되었습니다.")
        except Exception as e:
            print("오류 발생:", str(e))
            print("형식이 다른 파일이 끼어있습니다. 삭제 혹은 수정 후 다시 시도바랍니다")
    else:
        pass

combined()
a = input()
